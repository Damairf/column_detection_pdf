from sqlalchemy.orm import Session
from sqlalchemy import select
from fastapi import APIRouter, BackgroundTasks, UploadFile, File, Depends, HTTPException, status, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import List
import cv2
import os
import re
import shutil
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from fastapi.responses import StreamingResponse
from datetime import datetime, time

from database.database import get_db
from database import models
from services.auth import decode_access_token
from services.pdf_to_image import convert_pdf_to_images
from services.template_service import get_fields_from_db
from services.detection_pipeline import run_detection_pipeline

router   = APIRouter()
security = HTTPBearer()

# Helper
def get_current_user_id(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
) -> int:
    token   = credentials.credentials
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED,
                            detail="Token tidak valid atau sudah expired.")
    
    user_id = int(payload["sub"])
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED,
                            detail="Pengguna tidak ditemukan.")
    return user_id

def get_unique_filename(directory: str, filename: str) -> str:
    name, ext = os.path.splitext(filename)
    counter   = 1
    new_name  = filename
    while os.path.exists(os.path.join(directory, new_name)):
        new_name = f"{name}({counter}){ext}"
        counter += 1
    return new_name

def clean_filename(name: str) -> str:
    name = name.replace(" ", "_")
    name = re.sub(r"[^\w\-]", "", name)
    return name

class DokumenItem(BaseModel):
    nama_dokumen: str
    pdf_path: str

class SimpanDokumenRequest(BaseModel):
    id_spk: str
    id_template: int
    dokumen_list: List[DokumenItem]

class BatalUploadDokumenRequest(BaseModel):
    pdf_path: str

def _jalankan_deteksi(dokumen_id: int, db_url: str):

    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    engine  = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    db      = Session()

    try:
        dokumen = db.query(models.Dokumen).filter(
            models.Dokumen.id == dokumen_id
        ).first()

        if not dokumen:
            return

        template = db.query(models.Template).filter(
            models.Template.id == dokumen.id_template
        ).first()

        if not template:
            dokumen.status = "Error"
            db.commit()
            return

        image_folder = dokumen.path_dokumen

        if not image_folder or not os.path.isdir(image_folder):
            dokumen.status = "Error"
            db.commit()
            return

        import glob
        scan_images = sorted(
            glob.glob(os.path.join(image_folder, "*.png")) +
            glob.glob(os.path.join(image_folder, "*.jpg")) +
            glob.glob(os.path.join(image_folder, "*.jpeg"))
        )

        jml_halaman_dokumen  = len(scan_images)
        jml_halaman_template = template.jml_halaman or 1

        if jml_halaman_dokumen != jml_halaman_template:
            dokumen.status = "Error"
            db.commit()
            return

        db.query(models.HasilDeteksi).filter(
            models.HasilDeteksi.id_dokumen == dokumen_id
        ).delete(synchronize_session=False)
        db.commit()

        fields = get_fields_from_db(db, dokumen.id_template)

        if not fields:
            dokumen.status = "Error"
            db.commit()
            return

        pdf_path_template = template.path_template_pdf or ""
        nama_pdf          = os.path.basename(pdf_path_template)
        nama_tanpa_ext    = os.path.splitext(nama_pdf)[0]
        template_folder   = os.path.join("storage", "template", "images", nama_tanpa_ext)

        if not os.path.isdir(template_folder):
            base = os.path.join("storage", "template", "images")
            if os.path.isdir(base):
                for sub in sorted(os.listdir(base)):
                    full = os.path.join(base, sub)
                    if os.path.isdir(full):
                        template_folder = full
                        break

        results = []
        for halaman in range(1, jml_halaman_dokumen + 1):
            current_image_path = scan_images[halaman - 1]

            fields_per_page = {
                k: v for k, v in fields.items()
                if v["page"] == halaman
            }

            if not fields_per_page:
                continue

            try:
                page_result = run_detection_pipeline(
                    dokumen_image_folder = image_folder,
                    template_id          = dokumen.id_template,
                    fields               = fields_per_page,
                    template_image_base  = template_folder,
                    working_dir          = "storage/temp"
                )

                if isinstance(page_result, list):
                    for r in page_result:
                        r["page"] = halaman
                    results.extend(page_result)
                else:
                    page_result["page"] = halaman
                    results.append(page_result)

            except ValueError as e:
                dokumen.status = "Error"
                db.commit()
                return

        ada_kosong = False
        kriteria_count = 0
        benar_count = 0

        for page_result in results:
            halaman_result = page_result.get("page", 1)

            for field_name, status_kolom in page_result["results"].items():

                base_name = field_name.split("__hal")[0]

                kolom = db.query(models.KolomTemplate).filter(
                    models.KolomTemplate.id_template == dokumen.id_template,
                    models.KolomTemplate.nama_kolom  == base_name,
                    models.KolomTemplate.halaman     == halaman_result
                ).first()

                if kolom:
                    kriteria_count += 1
                    if status_kolom != "KOSONG":
                        benar_count += 1
                        
                    hasil = models.HasilDeteksi(
                        id_dokumen         = dokumen_id,
                        id_kolom_template  = kolom.id,
                        status             = status_kolom,
                    )
                    db.add(hasil)

                    if status_kolom == "KOSONG":
                        ada_kosong = True

        dokumen.status = "Salah" if ada_kosong else "Benar"
        db.commit()
        
        # Simpan ke tabel nilai
        if kriteria_count > 0:
            skor_val = round((benar_count / kriteria_count) * 100)
        else:
            skor_val = 0
            
        nilai_obj = models.Nilai(
            id_dokumen=dokumen_id,
            kriteria=float(kriteria_count),
            jml_benar=float(benar_count),
            skor=float(skor_val)
        )
        db.add(nilai_obj)
        db.commit()


    except Exception as e:
        print(f"[deteksi error] dokumen_id={dokumen_id}: {e}")
        try:
            dokumen = db.query(models.Dokumen).filter(
                models.Dokumen.id == dokumen_id
            ).first()
            if dokumen:
                dokumen.status = "Error"
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════════════════
# ROUTE
# ═══════════════════════════════════════════════════════════════════════

# ── GET /beranda/dokumen ──────────────────────────────────────────────
@router.get("/dokumen")
def get_dokumen_list(
    id_spk: str = Query(default=None),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")

    query = (
        db.query(
            models.Dokumen.id,
            models.Dokumen.nama_dokumen,
            models.Dokumen.status,
            models.Dokumen.path_dokumen,
            models.Dokumen.id_template,
            models.User.nama.label("pengunggah"),
            (
                select(models.Cabang.nama_cabang)
                .where(models.Cabang.id == models.User.id_cabang)
                .correlate(models.User)
                .scalar_subquery()
            ).label("cabang"),
            (
                select(models.Template.nama_template)
                .where(models.Template.id == models.Dokumen.id_template)
                .correlate(models.Dokumen)
                .scalar_subquery()
            ).label("nama_template"),
            models.Dokumen.created_at,
        )
        .outerjoin(models.User, models.Dokumen.id_user == models.User.id)
    )

    if user.role != "admin":
        query = query.filter(models.Dokumen.id_user == user_id)
        
    if id_spk:
        query = query.filter(models.Dokumen.id_spk == id_spk)

    results = query.order_by(models.Dokumen.created_at.desc()).all()

    return [
        {
            "id":            row.id,
            "nama_dokumen":  row.nama_dokumen,
            "pengunggah":    row.pengunggah,
            "cabang":        row.cabang,
            "status":        row.status,
            "path_dokumen":  row.path_dokumen,
            "id_template":   row.id_template,
            "nama_template": row.nama_template,
            "created_at":    row.created_at,
        }
        for row in results
    ]


# ── POST /beranda/upload-dokumen ──────────────────────────────────────
@router.post("/upload-dokumen")
async def upload_dokumen(
    file: UploadFile = File(...),
    user_id: int = Depends(get_current_user_id)
):

    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File harus berformat PDF.")

    pdf_dir   = "storage/dokumen/pdf"
    image_dir = "storage/dokumen/images"
    os.makedirs(pdf_dir,   exist_ok=True)
    os.makedirs(image_dir, exist_ok=True)

    unique_filename = get_unique_filename(pdf_dir, file.filename)
    pdf_path        = os.path.join(pdf_dir, unique_filename).replace("\\", "/")

    with open(pdf_path, "wb") as buf:
        shutil.copyfileobj(file.file, buf)

    try:
        nama_clean       = clean_filename(os.path.splitext(unique_filename)[0])
        image_output_dir = os.path.join(image_dir, nama_clean).replace("\\", "/")
        os.makedirs(image_output_dir, exist_ok=True)

        image_paths = convert_pdf_to_images(pdf_path, image_output_dir)

    except Exception as e:
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        raise HTTPException(status_code=400, detail=f"Gagal convert PDF ke image: {str(e)}")

    if not image_paths:
        raise HTTPException(status_code=400, detail="PDF tidak menghasilkan halaman.")

    img = cv2.imread(image_paths[0])
    if img is None:
        raise HTTPException(status_code=400, detail="Gagal membaca gambar hasil konversi.")

    height, width = img.shape[:2]

    return {
        "message":         "Dokumen berhasil diunggah",
        "nama_file":       unique_filename,
        "pdf_path":        pdf_path,
        "image_folder":    image_output_dir,
        "image_paths":     image_paths,
        "jml_halaman":     len(image_paths),
        "resolusi_width":  width,
        "resolusi_height": height,
    }


# ── DELETE /beranda/batal-upload-dokumen ──────────────────────────────
@router.delete("/batal-upload-dokumen")
def batal_upload_dokumen(
    data: BatalUploadDokumenRequest,
    user_id: int = Depends(get_current_user_id)
):
    pdf_path = data.pdf_path.strip()

    if not pdf_path.startswith("storage/dokumen/pdf/") or ".." in pdf_path:
        raise HTTPException(status_code=400, detail="Path tidak valid.")

    deleted = []
    if os.path.exists(pdf_path):
        os.remove(pdf_path)
        deleted.append(pdf_path)

    nama_file      = os.path.basename(pdf_path)
    nama_clean     = clean_filename(os.path.splitext(nama_file)[0])
    image_folder   = f"storage/dokumen/images/{nama_clean}"

    if os.path.isdir(image_folder):
        shutil.rmtree(image_folder)
        deleted.append(image_folder)

    return {"message": "File berhasil dihapus", "deleted": deleted}


# ── POST /beranda/simpan-dokumen ──────────────────────────────────────
@router.post("/simpan-dokumen")
def simpan_dokumen(
    data: SimpanDokumenRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):

    template = db.query(models.Template).filter(
        models.Template.id == data.id_template
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan.")

    from database.database import DATABASE_URL

    saved_ids = []

    for item in data.dokumen_list:
        nama_file  = os.path.basename(item.pdf_path)
        nama_clean = clean_filename(os.path.splitext(nama_file)[0])
        image_folder = f"storage/dokumen/images/{nama_clean}"

        dok = models.Dokumen(
            id_user      = user_id,
            nama_dokumen = item.nama_dokumen,
            status       = "Memuat",       
            path_dokumen = image_folder,        
            path_pdf     = item.pdf_path,       
            id_template  = data.id_template,
            id_spk       = data.id_spk,
        )
        db.add(dok)
        db.flush()  
        saved_ids.append(dok.id)

        background_tasks.add_task(_jalankan_deteksi, dok.id, DATABASE_URL)

    db.commit()

    return {
        "message":      f"{len(saved_ids)} dokumen berhasil disimpan dan sedang diproses.",
        "dokumen_ids":  saved_ids,
    }


# ── GET /beranda/dokumen/{dokumen_id} — detail dokumen + hasil deteksi ─
@router.get("/dokumen/{dokumen_id}")
def get_dokumen_detail(
    dokumen_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    
    query = db.query(models.Dokumen).filter(models.Dokumen.id == dokumen_id)
    if user and user.role != "admin":
        query = query.filter(models.Dokumen.id_user == user_id)
        
    dok = query.first()

    if not dok:
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan.")

    template = db.query(models.Template).filter(
        models.Template.id == dok.id_template
    ).first()
    nama_template = template.nama_template if template else None

    image_folder = dok.path_dokumen or ""
    folder_name  = os.path.basename(image_folder.rstrip("/\\"))

    pdf_path = ""
    pdf_dir  = "storage/dokumen/pdf"
    if folder_name and os.path.isdir(pdf_dir):
        for fname in os.listdir(pdf_dir):
            name_no_ext = clean_filename(os.path.splitext(fname)[0])
            if name_no_ext == folder_name:
                pdf_path = f"{pdf_dir}/{fname}".replace("\\", "/")
                break

    hasil_rows = (
        db.query(models.HasilDeteksi, models.KolomTemplate)
        .join(
            models.KolomTemplate,
            models.HasilDeteksi.id_kolom_template == models.KolomTemplate.id
        )
        .filter(models.HasilDeteksi.id_dokumen == dokumen_id)
        .order_by(models.KolomTemplate.halaman, models.KolomTemplate.id)
        .all()
    )

    hasil_list = [
        {
            "id_kolom":   kolom.id,
            "nama_kolom": kolom.nama_kolom,
            "halaman":    kolom.halaman,
            "x1":         kolom.x1,
            "y1":         kolom.y1,
            "x2":         kolom.x2,
            "y2":         kolom.y2,
            "status":     hasil.status,
        }
        for hasil, kolom in hasil_rows
    ]

    if not hasil_list and dok.status == "Memuat" and dok.id_template:
        koloms = db.query(models.KolomTemplate).filter(
            models.KolomTemplate.id_template == dok.id_template
        ).order_by(models.KolomTemplate.halaman, models.KolomTemplate.id).all()

        hasil_list = [
            {
                "id_kolom":   k.id,
                "nama_kolom": k.nama_kolom,
                "halaman":    k.halaman,
                "x1":         k.x1,
                "y1":         k.y1,
                "x2":         k.x2,
                "y2":         k.y2,
                "status":     "Memuat",
            }
            for k in koloms
        ]

    pengunggah = db.query(models.User).filter(models.User.id == dok.id_user).first()
    nama_pengunggah = pengunggah.nama if pengunggah else "Tidak diketahui"

    return {
        "dokumen": {
            "id":            dok.id,
            "nama_dokumen":  dok.nama_dokumen,
            "pengunggah":    nama_pengunggah,
            "nama_template": nama_template,
            "status":        dok.status,
            "path_pdf":      pdf_path,
            "path_dokumen":  dok.path_dokumen,
            "id_template":   dok.id_template,
            "created_at":    dok.created_at,
        },
        "hasil_deteksi": hasil_list,
    }

# ── DELETE /beranda/dokumen/{dokumen_id} — hapus dokumen ──────────────
@router.delete("/dokumen/{dokumen_id}")
def delete_dokumen(
    dokumen_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):

    user = db.query(models.User).filter(models.User.id == user_id).first()
    
    query = db.query(models.Dokumen).filter(models.Dokumen.id == dokumen_id)
    if user and user.role != "admin":
        query = query.filter(models.Dokumen.id_user == user_id)
        
    dok = query.first()

    if not dok:
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan.")

    deleted_files = []

    image_folder = dok.path_dokumen or ""
    folder_name  = os.path.basename(image_folder.rstrip("/\\"))

    if image_folder and os.path.isdir(image_folder):
        shutil.rmtree(image_folder)
        deleted_files.append(image_folder)

    pdf_dir = "storage/dokumen/pdf"
    if folder_name and os.path.isdir(pdf_dir):
        for fname in os.listdir(pdf_dir):
            name_no_ext = clean_filename(os.path.splitext(fname)[0])
            if name_no_ext == folder_name:
                pdf_full = os.path.join(pdf_dir, fname)
                os.remove(pdf_full)
                deleted_files.append(pdf_full)
                break

    db.query(models.HasilDeteksi).filter(
        models.HasilDeteksi.id_dokumen == dokumen_id
    ).delete(synchronize_session=False)

    db.query(models.Nilai).filter(
        models.Nilai.id_dokumen == dokumen_id
    ).delete(synchronize_session=False)

    db.delete(dok)
    db.commit()

    return {
        "message":"Dokumen berhasil dihapus",
        "deleted_files": deleted_files,
    }

# ── GET /beranda/download-dokumen — download dokumen (excel) ──────────
@router.get("/download-dokumen")
def download_dokumen(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d")
        ed = datetime.strptime(end_date, "%Y-%m-%d")
        ed = datetime.combine(ed, time.max)
    except ValueError:
        raise HTTPException(status_code=400, detail="Format tanggal tidak valid")

    query = db.query(models.Dokumen).filter(
        models.Dokumen.created_at >= sd,
        models.Dokumen.created_at <= ed
    )

    if user.role != "admin":
        query = query.filter(models.Dokumen.id_user == user_id)

    dokumen_list = query.order_by(models.Dokumen.id.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Dokumen"

    ws.column_dimensions['A'].width = 16.4
    ws.column_dimensions['B'].width = 52.1
    ws.column_dimensions['C'].width = 50.0
    ws.column_dimensions['D'].width = 25.0
    ws.column_dimensions['E'].width = 45.1
    ws.column_dimensions['F'].width = 20.1
    ws.column_dimensions['G'].width = 20.1

    if user.role == "admin":
        ws.column_dimensions['H'].width = 20.1
        headers = ["ID", "Nama Dokumen", "Pengunggah", "Cabang", "Nama Template", "Tanggal", "Status"]
    else:
        headers = ["ID", "Nama Dokumen", "Pengunggah", "Cabang", "Nama Template", "Tanggal"]

    ws.append(headers)

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for dok in dokumen_list:
        pengunggah_obj = db.query(models.User).filter(models.User.id == dok.id_user).first()
        nama_pengunggah = pengunggah_obj.nama if pengunggah_obj else "—"

        nama_cabang = "—"
        if pengunggah_obj and pengunggah_obj.id_cabang:
            cabang_obj = db.query(models.Cabang).filter(models.Cabang.id == pengunggah_obj.id_cabang).first()
            if cabang_obj:
                nama_cabang = cabang_obj.nama_cabang

        nama_template = "—"
        if dok.id_template:
            t = db.query(models.Template).filter(models.Template.id == dok.id_template).first()
            if t:
                nama_template = t.nama_template

        tgl = dok.created_at.strftime("%d/%m/%Y") if dok.created_at else "-"

        row = [
            dok.id,
            dok.nama_dokumen,
            nama_pengunggah,
            nama_cabang,
            nama_template,
            tgl,
        ]

        if user.role == "admin":
            row.append(dok.status)

        ws.append(row)

    data_font = Font(size=12)
    max_col_num = 7 if user.role == "admin" else 6
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col_num):
        for cell in row:
            cell.border = thin_border
            cell.font = data_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    try:
        sd_str = datetime.strptime(start_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        ed_str = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        sd_str, ed_str = start_date, end_date

    filename = f"Dokumen_{sd_str}_{ed_str}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )