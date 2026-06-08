from sqlalchemy.orm import Session
from sqlalchemy import select
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from database.database import get_db
from database import models
from routers.beranda_router import get_current_user_id
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime

router = APIRouter()

# ── GET /evaluasi/ ────────────────────────────────────────────────────
@router.get("/")
def get_evaluasi_list(
    start_date: Optional[str] = Query(default=None),
    end_date:   Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Akses ditolak")

    query = (
        db.query(
            models.Dokumen.id,
            models.Dokumen.nama_dokumen,
            models.User.nama.label("pengunggah"),
            (
                select(models.Cabang.nama_cabang)
                .where(models.Cabang.id == models.User.id_cabang)
                .correlate(models.User)
                .scalar_subquery()
            ).label("cabang"),
            models.SPK.id.label("nomor_spk"),
            models.SPK.nama_spk,
            models.SPK.tgl_retail,
            models.Nilai.kriteria,
            models.Nilai.jml_benar,
            models.Nilai.skor,
        )
        .outerjoin(models.User,  models.Dokumen.id_user  == models.User.id)
        .outerjoin(models.Nilai, models.Dokumen.id       == models.Nilai.id_dokumen)
        .outerjoin(models.SPK,   models.Dokumen.id_spk   == models.SPK.id)
    )

    if start_date:
        from datetime import date
        query = query.filter(models.SPK.tgl_retail >= date.fromisoformat(start_date))
    if end_date:
        from datetime import date
        query = query.filter(models.SPK.tgl_retail <= date.fromisoformat(end_date))

    results = query.order_by(models.Dokumen.created_at.desc()).all()

    return [
        {
            "id":           row.id,
            "nama_dokumen": row.nama_dokumen,
            "pengunggah":   row.pengunggah,
            "cabang":       row.cabang,
            "nomor_spk":    row.nomor_spk,
            "nama_spk":     row.nama_spk,
            "tgl_retail":   row.tgl_retail,
            "kriteria":     row.kriteria  if row.kriteria  is not None else 0,
            "jml_benar":    row.jml_benar if row.jml_benar is not None else 0,
            "skor":         row.skor      if row.skor      is not None else 0,
        }
        for row in results
    ]

# ── GET /evaluasi/ekspor ──────────────────────────────────────────────
@router.get("/ekspor")
def ekspor_evaluasi(
    cabang_ids:  Optional[List[int]] = Query(default=None),
    start_date:  Optional[str]       = Query(default=None),
    end_date:    Optional[str]       = Query(default=None),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Akses ditolak")

    query = (
        db.query(
            models.Dokumen.id,
            models.Dokumen.nama_dokumen,
            models.User.nama.label("pengunggah"),
            (
                select(models.Cabang.nama_cabang)
                .where(models.Cabang.id == models.User.id_cabang)
                .correlate(models.User)
                .scalar_subquery()
            ).label("cabang"),
            models.SPK.id.label("nomor_spk"),
            models.SPK.nama_spk,
            models.SPK.tgl_retail,
            models.Nilai.kriteria,
            models.Nilai.jml_benar,
            models.Nilai.skor,
        )
        .outerjoin(models.User, models.Dokumen.id_user == models.User.id)
        .outerjoin(models.Nilai, models.Dokumen.id == models.Nilai.id_dokumen)
        .outerjoin(models.SPK, models.Dokumen.id_spk == models.SPK.id)
    )

    if cabang_ids:
        query = query.filter(models.User.id_cabang.in_(cabang_ids))

    if start_date:
        from datetime import date
        query = query.filter(models.SPK.tgl_retail >= date.fromisoformat(start_date))
    if end_date:
        from datetime import date
        query = query.filter(models.SPK.tgl_retail <= date.fromisoformat(end_date))

    results = query.order_by(models.Dokumen.id.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Evaluasi"

    ws.column_dimensions['A'].width = 16.4
    ws.column_dimensions['B'].width = 52.1
    ws.column_dimensions['C'].width = 30.0
    ws.column_dimensions['D'].width = 25.0
    ws.column_dimensions['E'].width = 20.0
    ws.column_dimensions['F'].width = 45.1
    ws.column_dimensions['G'].width = 20.1
    ws.column_dimensions['H'].width = 18.1
    ws.column_dimensions['I'].width = 18.1
    ws.column_dimensions['J'].width = 18.1

    headers = ["ID", "Nama Dokumen", "Pengunggah", "Cabang", "Nomor SPK", "Nama SPK", "Tanggal Retail", "Kriteria", "Benar", "Skor"]
    ws.append(headers)

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for row in results:
        tgl = row.tgl_retail.strftime("%d/%m/%Y") if row.tgl_retail else "—"
        kriteria  = int(row.kriteria)  if row.kriteria  is not None else 0
        jml_benar = int(row.jml_benar) if row.jml_benar is not None else 0
        skor      = int(row.skor)      if row.skor      is not None else 0

        ws.append([
            row.id,
            row.nama_dokumen or "—",
            row.pengunggah   or "—",
            row.cabang       or "—",
            row.nomor_spk    or "—",
            row.nama_spk     or "—",
            tgl,
            kriteria,
            f"{jml_benar}/{kriteria}",
            skor / 100,
        ])

    data_font = Font(size=12)
    for data_row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in data_row:
            cell.border = thin_border
            cell.font = data_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    tgl_mulai   = date.fromisoformat(start_date).strftime("%d-%m-%Y") if start_date else "tanpa-tanggal"
    tgl_selesai = date.fromisoformat(end_date).strftime("%d-%m-%Y")   if end_date   else "tanpa-tanggal"
    filename = f"Evaluasi-Dokumen_{tgl_mulai}_{tgl_selesai}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )